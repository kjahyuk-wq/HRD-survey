#!/usr/bin/env python3
"""신규자과정 좌석 조회 페이지 빌드 스크립트.

명단·좌석배치 엑셀을 읽어 개인정보를 암호화한 정적 HTML(seat-lookup/index.html)을 생성한다.
페이지 소스에는 이름·전화번호 원본이 들어가지 않는다:
  - 조회 키: PBKDF2-SHA256(이름|뒤4자리, salt, 200k회) 앞 16바이트(hex)
  - 레코드: 같은 파생값의 뒤 32바이트를 AES-256-GCM 키로 사용해 암호화

사용법:  python3 build_seat_lookup.py [엑셀비밀번호]
"""
import sys, os, io, json, base64, hashlib, secrets
import openpyxl
from cryptography.hazmat.primitives.ciphers.aead import AESGCM

BASE = os.path.dirname(os.path.abspath(__file__))
ROSTER_XLSX = os.path.join(BASE, "2026년_제2기_신규자과정_분임편성_및_교육생명단.xlsx")
SEATS_XLSX = os.path.join(BASE, "대강당 1층 441명+별도2_좌석배치_띄어앉기.xlsx")
TEMPLATE = os.path.join(BASE, "seat-lookup.template.html")
OUT_DIR = os.path.join(BASE, "seat-lookup")

SALT = "djhrd-2026-2-newbie-v1"
ITER = 200_000

# ── 설정 ─────────────────────────────────────────────────────────
# 거동불편 별도책상(후방) 배정 교번 2명. 확정되면 채우고 재실행.
# 예: CARE = [122, 377]
CARE = []

# 체육대회 조 매핑. {교번: "청팀"} 또는 분임→조 규칙이 정해지면 team_of()를 수정.
def team_of(row):
    return None  # 아직 미정 → 페이지에 "추후 안내" 표시
# ─────────────────────────────────────────────────────────────────


def load_roster(password):
    with open(ROSTER_XLSX, "rb") as fp:
        head = fp.read(8)
        fp.seek(0)
        if head[:4] == b"PK\x03\x04":  # 암호 해제된 파일
            wb = openpyxl.load_workbook(fp, data_only=True)
        else:
            import msoffcrypto
            off = msoffcrypto.OfficeFile(fp)
            off.load_key(password=password)
            buf = io.BytesIO()
            off.decrypt(buf)
            wb = openpyxl.load_workbook(buf, data_only=True)
    ws = wb["분임편성 명단"]
    people = []
    for r in range(5, ws.max_row + 1):
        eid = ws.cell(r, 2).value  # 교번
        if eid is None:
            continue
        people.append({
            "eid": int(eid),
            "name": str(ws.cell(r, 5).value).strip(),
            "phone": str(ws.cell(r, 9).value).strip(),
            "bunim": int(ws.cell(r, 14).value),  # 분임(집계용)
        })
    people.sort(key=lambda p: p["eid"])
    return people


def load_seats():
    wb = openpyxl.load_workbook(SEATS_XLSX, data_only=True)
    ws = wb["대강당 (1층)"]
    grid = []
    for r in range(7, 22):  # 1~15열
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)) and 1 <= v <= 441:
                grid.append([int(v), r - 6, c])
    assert len(grid) == 441, f"좌석 수 이상: {len(grid)}"
    return sorted(grid)


def assign_seats(people):
    """교번 순서대로 좌석 1~441 배정. CARE 교번은 별도책상(care), 초과 인원은 미정."""
    next_seat = 1
    for p in people:
        if p["eid"] in CARE:
            p["seat"], p["care"] = None, True
        elif next_seat <= 441:
            p["seat"], p["care"] = next_seat, False
            next_seat += 1
        else:
            p["seat"], p["care"] = None, False
    return people


def encrypt_records(people):
    data = {}
    for p in people:
        material = f'{p["name"].replace(" ", "")}|{p["phone"][-4:]}'.encode()
        dk = hashlib.pbkdf2_hmac("sha256", material, SALT.encode(), ITER, dklen=48)
        lookup_id, aes_key = dk[:16].hex(), dk[16:48]
        rec = {"n": p["name"], "e": p["eid"], "b": p["bunim"],
               "t": team_of(p), "s": p["seat"]}
        if p["care"]:
            rec["c"] = 1
        iv = secrets.token_bytes(12)
        ct = AESGCM(aes_key).encrypt(iv, json.dumps(rec, ensure_ascii=False).encode(), None)
        assert lookup_id not in data, f'조회키 충돌: {p["name"]}'
        data[lookup_id] = base64.b64encode(iv + ct).decode()
    return data


def main():
    password = sys.argv[1] if len(sys.argv) > 1 else "66640"
    people = assign_seats(load_roster(password))
    grid = load_seats()
    data = encrypt_records(people)

    with open(TEMPLATE, encoding="utf-8") as f:
        html = f.read()
    html = (html
            .replace("__DATA__", json.dumps(data, separators=(",", ":")))
            .replace("__GRID__", json.dumps(grid, separators=(",", ":")))
            .replace("__SALT__", SALT)
            .replace("__ITER__", str(ITER)))

    os.makedirs(OUT_DIR, exist_ok=True)
    out = os.path.join(OUT_DIR, "index.html")
    with open(out, "w", encoding="utf-8") as f:
        f.write(html)

    unseated = [p["eid"] for p in people if p["seat"] is None and not p["care"]]
    print(f"완료: {out} ({os.path.getsize(out)//1024} KB)")
    print(f"인원 {len(people)}명 / 좌석배정 {sum(1 for p in people if p['seat'])}명 "
          f"/ 별도책상 {len(CARE)}명 / 미정 {unseated}")


if __name__ == "__main__":
    main()
