# 삭제된 정식 만족도 과정 2건을 Firestore 에 '종료된 과정'으로 복원
#   1) 예산 및 회계(1기)      — 원시 개별응답 export 를 그대로 디코딩 (완전 복원)
#   2) 헌법과 생활법률 1기    — 통계 export 의 문항별 분포와 정확히 일치하도록 응답 재구성
# 사용: python3 restore.py [--write]   (--write 없으면 검증만)
import json, re, sys, random, urllib.request, urllib.parse
import openpyxl
from pathlib import Path

SC = Path(__file__).parent
PROJECT = 'hrd-data'
BASE = f'https://firestore.googleapis.com/v1/projects/{PROJECT}/databases/(default)/documents'

DEMO = {
    'q11': ['시 본청', '시 사업소', '구', '동', '기타'],
    'q12': ['5급', '6급', '7급', '8급', '9급', '기타'],
    'q13': ['행정직', '기술직', '연구직', '관리운영직', '기타'],
    'q14': ['20대', '30대', '40대', '50대'],
    'q15': ['남', '여'],
    'q16': ['업무능력 개발', '교육이수 점수 취득', '심신의 재충전', '자기개발', '기타'],
}
DEMO_KEYS = list(DEMO.keys())

def load_stats_book(path):
    """만족도결과 xlsx → {name, n, qdists(9x5), inst:[(edu,name,dist5)], demo:{key:[counts]}, comments:[[4str]]}"""
    wb = openpyxl.load_workbook(path)
    s1 = [list(r) for r in wb['만족도 통계'].iter_rows(values_only=True)]
    name = str(s1[0][1]).strip()
    n = int(re.sub(r'\D', '', str(s1[1][1])))
    qdists, inst = [], []
    for row in s1:
        c0 = str(row[0] or '')
        if re.match(r'^Q[1-9]\.', c0):
            qdists.append([int(row[i]) for i in range(3, 8)])
        m = re.match(r'^\[(.+)\] (\S+) 강사$', c0)
        if m:
            inst.append((m.group(1), m.group(2), [int(row[i]) for i in range(3, 8)]))
    assert len(qdists) == 9, qdists
    demo = {}
    cur = None
    for row in wb['응답자 특성'].iter_rows(values_only=True):
        c0 = str(row[0] or '').strip()
        m = re.match(r'^(Q1[1-6])\.', c0)
        if m:
            cur = 'q' + m.group(1)[1:]
            demo[cur] = []
        elif cur and c0 and c0 != '합계' and row[1] is not None:
            demo[cur].append(int(row[1]))
    for k in DEMO_KEYS:
        assert len(demo[k]) == len(DEMO[k]), (k, demo[k])
        assert sum(demo[k]) == n, (k, demo[k], n)
    comments = []
    for row in wb['주관식 의견'].iter_rows(values_only=True):
        if str(row[0] or '').strip().isdigit():
            comments.append([str(c or '') for c in row[1:5]])
    return dict(name=name, n=n, qdists=qdists, inst=inst, demo=demo, comments=comments)

def synthesize(stats, seed=42):
    """문항별 주변분포가 통계와 정확히 일치하는 응답 n건 재구성 (조합은 셔플)"""
    rng = random.Random(seed)
    n = stats['n']
    docs = [dict() for _ in range(n)]
    def scores_from(dist):
        vals = [s for s, c in enumerate(dist, start=1) for _ in range(c)]
        assert len(vals) == n
        rng.shuffle(vals)
        return vals
    for qi, dist in enumerate(stats['qdists'], start=1):
        for d, v in zip(docs, scores_from(dist)):
            d[f'q{qi}'] = v
    for d in docs:
        d['instructors'] = {}
    for edu, iname, dist in stats['inst']:
        key = f'{edu}__{iname}'
        for d, v in zip(docs, scores_from(dist)):
            d['instructors'][key] = v
    for k in DEMO_KEYS:
        opts = [o for o, c in zip(DEMO[k], stats['demo'][k]) for _ in range(c)]
        rng.shuffle(opts)
        for d, v in zip(docs, opts):
            d[k] = v
    attach_comments(docs, stats['comments'])
    return docs

def attach_comments(docs, comments):
    for d in docs:
        d.setdefault('q10_comment', '')
        d.setdefault('comment1', '')
        d.setdefault('comment2', '')
        d.setdefault('comment3', '')
    for i, (q10, c1, c2, c3) in enumerate(comments):
        docs[i]['q10_comment'], docs[i]['comment1'], docs[i]['comment2'], docs[i]['comment3'] = q10, c1, c2, c3

def decode_raw(raw_path, inst_keys):
    """설문결과(원시) xlsx → 응답 docs. export 인코딩: 표시값 = 6 - 저장점수, 인구통계 = 보기순번_0"""
    wb = openpyxl.load_workbook(raw_path)
    rows = [list(r) for r in wb['객관식'].iter_rows(values_only=True)][1:]
    docs = []
    for row in rows:
        d = {}
        for qi in range(1, 10):
            v = int(row[qi])
            assert 1 <= v <= 5
            d[f'q{qi}'] = 6 - v
        for j, k in enumerate(DEMO_KEYS):
            cell = str(row[11 + j] or '')
            m = re.match(r'^(\d+)_0$', cell)
            assert m, cell
            d[k] = DEMO[k][int(m.group(1)) - 1]
        d['instructors'] = {}
        for j, key in enumerate(inst_keys):
            cell = str(row[17 + j] or '')
            m = re.match(r'^(\d+)_0$', cell)
            assert m, cell
            d['instructors'][key] = 6 - int(m.group(1))
        docs.append(d)
    comments = []
    for row in wb['주관식'].iter_rows(values_only=True):
        if str(row[0] or '').strip().isdigit():
            comments.append([str(c or '') for c in row[1:5]])
    attach_comments(docs, comments)
    return docs

def verify(docs, stats):
    n = stats['n']
    assert len(docs) == n, (len(docs), n)
    for qi, want in enumerate(stats['qdists'], start=1):
        got = [0] * 5
        for d in docs:
            got[d[f'q{qi}'] - 1] += 1
        assert got == want, (f'q{qi}', got, want)
    for edu, iname, want in stats['inst']:
        key = f'{edu}__{iname}'
        got = [0] * 5
        for d in docs:
            got[d['instructors'][key] - 1] += 1
        assert got == want, (key, got, want)
    for k in DEMO_KEYS:
        got = [sum(1 for d in docs if d[k] == o) for o in DEMO[k]]
        assert got == stats['demo'][k], (k, got, stats['demo'][k])
    print(f"  ✓ {stats['name']}: 응답 {n}건 — 문항 9개 분포/강사 {len(stats['inst'])}명 분포/인구통계 6문항 모두 통계 export 와 정확히 일치")

# ── Firestore REST ──────────────────────────────
def fs_value(v):
    if isinstance(v, bool): return {'booleanValue': v}
    if isinstance(v, int): return {'integerValue': str(v)}
    if isinstance(v, str): return {'stringValue': v}
    if isinstance(v, dict): return {'mapValue': {'fields': {k: fs_value(x) for k, x in v.items()}}}
    if isinstance(v, tuple) and v[0] == 'ts': return {'timestampValue': v[1]}
    raise TypeError(v)

def get_token():
    cfg = json.loads((Path.home() / '.config/configstore/firebase-tools.json').read_text())
    body = urllib.parse.urlencode({
        'grant_type': 'refresh_token',
        'refresh_token': cfg['tokens']['refresh_token'],
        'client_id': '563584335869-fgrhgmd47bqnekij5i8b5pr03ho849e6.apps.googleusercontent.com',
        'client_secret': 'j9iVZfS8kkCEFUPaAeJV0sAi',
    }).encode()
    req = urllib.request.Request('https://oauth2.googleapis.com/token', data=body)
    return json.loads(urllib.request.urlopen(req).read())['access_token']

def commit(token, writes):
    req = urllib.request.Request(
        f'https://firestore.googleapis.com/v1/projects/{PROJECT}/databases/(default)/documents:commit',
        data=json.dumps({'writes': writes}).encode(),
        headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'},
    )
    return json.loads(urllib.request.urlopen(req).read())

def auto_id(rng):
    chars = 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789'
    return ''.join(rng.choice(chars) for _ in range(20))

def build_writes(course, rng):
    doc_prefix = f'projects/{PROJECT}/databases/(default)/documents'
    cid = auto_id(rng)
    writes = []
    def put(path, fields):
        writes.append({'update': {'name': f'{doc_prefix}/{path}',
                                  'fields': {k: fs_value(v) for k, v in fields.items()}}})
    end_ts = ('ts', f"{course['endDate']}T08:00:00Z")  # 종료일 17:00 KST
    put(f'courses/{cid}', {
        'name': course['name'], 'active': False, 'type': 'standard',
        'startDate': course['startDate'], 'endDate': course['endDate'],
        'restored': True, 'restoredAt': ('ts', course['restoredAt']),
        'restoredSource': course['source'],
    })
    for i, (edu, iname, _d) in enumerate(course['inst']):
        put(f'courses/{cid}/instructors/{auto_id(rng)}',
            {'name': iname, 'education': edu, 'order': (i + 1) * 10, 'createdAt': end_ts})
    for i in range(course['n']):
        put(f'courses/{cid}/students/{auto_id(rng)}',
            {'name': '', 'empNo': str(i + 1), 'completed': True, 'completedAt': end_ts})
    for i, d in enumerate(course['docs']):
        fields = dict(d)
        fields['name'] = ''
        fields['empNo'] = str(i + 1)
        fields['course'] = course['name']
        fields['submittedAt'] = ('ts', f"{course['endDate']}T{6 + (i * 3) // 60:02d}:{(i * 3) % 60:02d}:00Z")
        fields['restored'] = True
        put(f'courses/{cid}/responses/{auto_id(rng)}', fields)
    return cid, writes

def main():
    write = '--write' in sys.argv
    print('── 1) 예산 및 회계(1기): 원시 응답 디코딩 ──')
    stats_b = load_stats_book(SC / 'restore-2.xlsx')
    inst_keys_b = [f'{e}__{n}' for e, n, _ in stats_b['inst']]
    docs_b = decode_raw(SC / 'raw-budget.xlsx', inst_keys_b)
    verify(docs_b, stats_b)

    print('── 2) 헌법과 생활법률 1기: 분포 기반 재구성 ──')
    stats_l = load_stats_book(SC / 'restore-1.xlsx')
    docs_l = synthesize(stats_l)
    verify(docs_l, stats_l)

    courses = [
        dict(name=stats_l['name'], n=stats_l['n'], inst=stats_l['inst'], docs=docs_l,
             startDate='2026-03-16', endDate='2026-03-18',
             source='로컬 export 통계(만족도결과 2026-03-20)에서 분포 재구성',
             restoredAt='2026-07-13T00:00:00Z'),
        dict(name=stats_b['name'], n=stats_b['n'], inst=stats_b['inst'], docs=docs_b,
             startDate='2026-03-23', endDate='2026-03-25',
             source='로컬 export 원시응답(설문결과 2026-03-25) 완전 복원',
             restoredAt='2026-07-13T00:00:00Z'),
    ]
    (SC / 'restore-data.json').write_text(json.dumps(
        [{k: v for k, v in c.items() if k != 'inst'} | {'inst': [(e, n) for e, n, _ in c['inst']]} for c in courses],
        ensure_ascii=False, indent=1, default=str))
    if not write:
        print('검증만 완료 (--write 로 실제 기록)')
        return
    token = get_token()
    rng = random.Random(7)
    for c in courses:
        cid, writes = build_writes(c, rng)
        res = commit(token, writes)
        print(f"  ✓ 기록 완료: {c['name']} → courses/{cid} (문서 {len(writes)}개)")

if __name__ == '__main__':
    main()
